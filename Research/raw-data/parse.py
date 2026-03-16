import openpyxl, json, os

NAME_TO_ISO3 = {
    "Afghanistan":"AFG","Albania":"ALB","Algeria":"DZA","Angola":"AGO",
    "Argentina":"ARG","Armenia":"ARM","Australia":"AUS","Austria":"AUT",
    "Azerbaijan":"AZE","Bahrain":"BHR","Bangladesh":"BGD","Belarus":"BLR",
    "Belgium":"BEL","Bolivia":"BOL","Bosnia and Herzegovina":"BIH","Botswana":"BWA",
    "Brazil":"BRA","Brunei Darussalam":"BRN","Bulgaria":"BGR","Burkina Faso":"BFA",
    "Cambodia":"KHM","Cameroon":"CMR","Canada":"CAN","Chad":"TCD","Chile":"CHL",
    "China, P.R.":"CHN","China":"CHN","Colombia":"COL","Congo, Dem. Rep.":"COD",
    "Congo, Republic":"COG","Costa Rica":"CRI","Croatia":"HRV",
    "Cuba":"CUB","Cyprus":"CYP","Czechia":"CZE","Czech Republic":"CZE",
    "Denmark":"DNK","Dominican Republic":"DOM","Ecuador":"ECU","Egypt":"EGY",
    "El Salvador":"SLV","Estonia":"EST","Ethiopia":"ETH","Finland":"FIN",
    "France":"FRA","Gabon":"GAB","Georgia":"GEO","Germany":"DEU","Ghana":"GHA",
    "Greece":"GRC","Guatemala":"GTM","Honduras":"HND","Hungary":"HUN",
    "India":"IND","Indonesia":"IDN","Iran":"IRN","Iraq":"IRQ","Ireland":"IRL",
    "Israel":"ISR","Italy":"ITA","Jamaica":"JAM","Japan":"JPN","Jordan":"JOR",
    "Kazakhstan":"KAZ","Kenya":"KEN","Korea, North":"PRK","Korea, South":"KOR",
    "Kuwait":"KWT","Kyrgyzstan":"KGZ","Latvia":"LVA","Lebanon":"LBN",
    "Libya":"LBY","Lithuania":"LTU","Luxembourg":"LUX","Malaysia":"MYS",
    "Mali":"MLI","Mexico":"MEX","Moldova":"MDA","Mongolia":"MNG","Morocco":"MAR",
    "Mozambique":"MOZ","Myanmar":"MMR","Namibia":"NAM","Nepal":"NPL",
    "Netherlands":"NLD","New Zealand":"NZL","Nicaragua":"NIC","Niger":"NER",
    "Nigeria":"NGA","North Macedonia":"MKD","Norway":"NOR","Oman":"OMN",
    "Pakistan":"PAK","Panama":"PAN","Paraguay":"PRY","Peru":"PER","Philippines":"PHL",
    "Poland":"POL","Portugal":"PRT","Qatar":"QAT","Romania":"ROU","Russia":"RUS",
    "Rwanda":"RWA","Saudi Arabia":"SAU","Senegal":"SEN","Serbia":"SRB",
    "Singapore":"SGP","Slovakia":"SVK","Slovenia":"SVN","Somalia":"SOM",
    "South Africa":"ZAF","South Sudan":"SSD","Spain":"ESP","Sri Lanka":"LKA",
    "Sudan":"SDN","Sweden":"SWE","Switzerland":"CHE","Syria":"SYR","Taiwan":"TWN",
    "Tajikistan":"TJK","Tanzania":"TZA","Thailand":"THA","Timor-Leste":"TLS",
    "Togo":"TGO","Trinidad and Tobago":"TTO","Tunisia":"TUN",
    "Turkey":"TUR","Turkmenistan":"TKM","Uganda":"UGA","Ukraine":"UKR",
    "United Arab Emirates":"ARE","United Kingdom":"GBR","United States":"USA",
    "Uruguay":"URY","Uzbekistan":"UZB","Venezuela":"VEN","Viet Nam":"VNM",
    "Vietnam":"VNM","Yemen":"YEM","Zambia":"ZMB","Zimbabwe":"ZWE",
    "Montenegro":"MNE","Kosovo":"XKX","North Korea":"PRK","South Korea":"KOR",
    "Czechia*":"CZE","Denmark*":"DNK","Estonia*":"EST","France*":"FRA",
    "Germany*":"DEU","Greece*":"GRC","Hungary*":"HUN","Luxembourg*":"LUX",
    "Montenegro*":"MNE","Netherlands*":"NLD","North Macedonia*":"MKD",
    "Poland*":"POL","Romania*":"ROU","Slovakia*":"SVK","Spain*":"ESP",
    "Turkey*":"TUR","United Kingdom*":"GBR","United States*":"USA",
    "Cote d'Ivoire":"CIV","Côte d'Ivoire":"CIV","Türkiye":"TUR",
}

BASE = "/Users/matteo/Documents/GitHub/hmwwct"

# ─── 1. SIPRI Military Expenditure (Current USD millions) ────────────────────
wb = openpyxl.load_workbook(f'{BASE}/Research/raw-data/sipri-milex.xlsx', read_only=True, data_only=True)
ws = wb['Current US$']
rows = list(ws.iter_rows(values_only=True))

header = rows[5]
year_indices = {}
for i in range(2, len(header)):
    if isinstance(header[i], int):
        year_indices[str(header[i])] = i
recent_years = [y for y in year_indices if int(y) >= 2010]

SECTION_HEADERS = {
    'Africa','North Africa','sub-Saharan Africa','Americas',
    'Central America and the Caribbean','North America','South America',
    'Asia and Oceania','Central Asia','East Asia','Oceania','South Asia',
    'South East Asia','Europe','Central Europe','Eastern Europe',
    'Western Europe','Middle East','Middle East and North Africa',
}

sipri_data = {}
for row in rows[6:]:
    name = row[0]
    if not name or not isinstance(name, str):
        continue
    clean = name.strip()
    if clean in SECTION_HEADERS or clean == '':
        continue
    iso = NAME_TO_ISO3.get(clean)
    if not iso:
        continue
    country_years = {}
    for y in recent_years:
        val = row[year_indices[y]]
        if isinstance(val, (int, float)):
            country_years[y] = round(float(val), 1)
    if country_years:
        sipri_data[iso] = {"name": clean, "milexMillionUSD": country_years}

wb.close()
print(f"SIPRI: {len(sipri_data)} countries parsed")

# ─── 2. NATO Defence Data ────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(f'{BASE}/Research/raw-data/nato-def-exp.xlsx', read_only=True, data_only=True)

def parse_nato_table(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = rows[3]
    year_cols = []
    for i in range(1, len(header)):
        val = header[i]
        if val is None:
            continue
        s = str(val).replace('e','').strip()
        if s.isdigit() and 2010 <= int(s) <= 2030:
            year_cols.append((s, i))
    result = {}
    for row in rows[4:]:
        name = row[0]
        if not name or not isinstance(name, str):
            continue
        clean = name.strip()
        iso = NAME_TO_ISO3.get(clean)
        if not iso:
            continue
        d = {}
        for y, i in year_cols:
            val = row[i]
            if isinstance(val, (int, float)):
                d[y] = round(float(val), 2)
        if d:
            result[iso] = {"name": clean, "values": d}
    return result

nato_spend_usd    = parse_nato_table(wb2['Table 2'])   # defence spend USD millions (current)
nato_equipment_pct = parse_nato_table(wb2['Table 8a']) # equipment % of total spend
nato_gdp_pct       = parse_nato_table(wb2['Table 3'])  # defence as % of GDP
wb2.close()
print(f"NATO spend USD: {len(nato_spend_usd)} | equipment %: {len(nato_equipment_pct)} | gdp %: {len(nato_gdp_pct)}")

# ─── 3. Bruegel FMS (US Foreign Military Sales) ──────────────────────────────
wb3 = openpyxl.load_workbook(f'{BASE}/Research/raw-data/bruegel-fms.xlsx', read_only=True, data_only=True)
ws3 = wb3['MAINDATA']
rows3 = list(ws3.iter_rows(values_only=True))
header3 = rows3[0]
col = {v: i for i, v in enumerate(header3) if v is not None}

bruegel_by_country = {}
bruegel_by_category = {}

for row in rows3[1:]:
    country = row[col['country']]
    val = row[col['financial_value_2024_prices']]
    eq_type = row[col['general_item_type']]
    domain = row[col['military_domain']]
    year = row[col['year']]
    if not isinstance(val, (int, float)) or not country:
        continue
    val_m = round(float(val) * 1000, 2)  # billions → millions USD 2024

    if country not in bruegel_by_country:
        bruegel_by_country[country] = {"totalMillionUSD2024": 0.0, "byYear": {}, "byCategory": {}}
    bruegel_by_country[country]["totalMillionUSD2024"] = round(
        bruegel_by_country[country]["totalMillionUSD2024"] + val_m, 2)
    yr = str(year)
    bruegel_by_country[country]["byYear"][yr] = round(
        bruegel_by_country[country]["byYear"].get(yr, 0.0) + val_m, 2)
    if eq_type:
        bruegel_by_country[country]["byCategory"][eq_type] = round(
            bruegel_by_country[country]["byCategory"].get(eq_type, 0.0) + val_m, 2)

    if eq_type:
        bruegel_by_category[eq_type] = round(bruegel_by_category.get(eq_type, 0.0) + val_m, 2)

wb3.close()
print(f"Bruegel: {len(bruegel_by_country)} recipient countries")

# ─── 4. Write JSON files ─────────────────────────────────────────────────────
out_dir = f'{BASE}/src/lib/data/armaments'
os.makedirs(out_dir, exist_ok=True)

with open(f'{out_dir}/sipri-milex.json', 'w') as f:
    json.dump({
        "source": "SIPRI Military Expenditure Database 2024",
        "description": "Military expenditure by country in millions of current USD, 2010-2024",
        "unit": "millions USD (current prices)",
        "countries": sipri_data
    }, f, indent=2)

with open(f'{out_dir}/nato-defence.json', 'w') as f:
    json.dump({
        "source": "NATO Defence Expenditure 2025",
        "description": "NATO member defence data: total spend (USD millions), equipment % of spend, defence % of GDP",
        "spendMillionUSD": nato_spend_usd,
        "equipmentPct": nato_equipment_pct,
        "gdpPct": nato_gdp_pct
    }, f, indent=2)

with open(f'{out_dir}/bruegel-fms.json', 'w') as f:
    json.dump({
        "source": "Bruegel US Foreign Military Sales Database (Apr 2008 – Sep 2025)",
        "description": "US arms sold to each recipient country, millions USD at 2024 constant prices",
        "unit": "millions USD (2024 constant prices)",
        "recipientCountries": bruegel_by_country,
        "byEquipmentCategory": dict(sorted(bruegel_by_category.items(), key=lambda x: -x[1])[:60])
    }, f, indent=2)

print("\n✓ Written:")
for fn in ['sipri-milex.json', 'nato-defence.json', 'bruegel-fms.json']:
    size = os.path.getsize(f'{out_dir}/{fn}')
    print(f"  {fn}: {size // 1024} KB")
